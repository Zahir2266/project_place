from django.shortcuts import render
from rest_framework import viewsets, permissions, filters as drf_filters
from .models import Location, Event
from .serializers import LocationSerializer, EventSerializer
from .filters import EventFilter
from drf_spectacular.utils import extend_schema, extend_schema_view

# Экспорт таблиц
import openpyxl
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils.dateparse import parse_datetime

@extend_schema(tags=['Места проведения'])
@extend_schema_view(
    list=extend_schema(summary="Список мест"),
    create=extend_schema(
        summary="Создать место",
        request=LocationSerializer,
    ),
    retrieve=extend_schema(summary="Просмотр одного места", description="Доступно только администратору"),
    update=extend_schema(summary="Изменить место", description="Доступно только администратору"),
    partial_update=extend_schema(summary="Изменить место (частично)", description="Доступно только администратору"),
    destroy=extend_schema(summary="Удалить место", description="Доступно только администратору"),
)
class LocationViewSet(viewsets.ModelViewSet):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer

    # Ограничение на доступа только суперюзу
    permission_classes = [permissions.IsAdminUser]



@extend_schema(tags=['Мероприятия'])
@extend_schema_view(
    list=extend_schema(
        summary="Получить список всех мероприятий", 
        description="Обычные пользователи видят только опубликованные мероприятия. Суперпользователи видят всё."
    ),
    retrieve=extend_schema(
        summary="Детальная информация о мероприятии",
        description="Позволяет увидеть погоду, рейтинг и полный список изображений."
    ),
    # create -  описание загрузки файлов
    create=extend_schema(
        summary="Создать новое мероприятие",
        description="Доступно только администратору. Поле uploaded_images позволяет выбрать несколько файлов.",
        
        # Ручное переопределение схемы запроса для поддержки multipart/form-data (загрузка файлов)
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'title': {'type': 'string'},
                    'description': {'type': 'string'},
                    'location': {'type': 'integer'},
                    'start_date': {'type': 'string', 'format': 'date-time'},
                    'end_date': {'type': 'string', 'format': 'date-time'},
                    'rating': {'type': 'integer', 'minimum': 0, 'maximum': 25},
                    'status': {'type': 'string', 'enum': ['draft', 'published']},
                    'uploaded_images': {
                        'type': 'array',
                        'items': {'type': 'string', 'format': 'binary'}
                    }
                },
                'required': ['title', 'location', 'start_date', 'end_date']
            }
        },
    ),
    update=extend_schema(summary="Редактировать мероприятие", description="Только для администратора"),
    partial_update=extend_schema(summary="Редактировать мероприятие (частично)", description="Только для администратора"),
    destroy=extend_schema(summary="Удалить мероприятие", description="Только для администратора"),
)
class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all()
    serializer_class = EventSerializer

    filterset_class = EventFilter
    # Поиск по названию или месту
    search_fields = ['title', 'location__name']  
    # Сортировка
    ordering_fields = ['title', 'start_date', 'end_date'] 
     # Сортировка по умолчанию
    ordering = ['title']
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAdminUser()]
        return [permissions.AllowAny()]

    def get_queryset(self):
        user = self.request.user
        queryset = Event.objects.all()
        
        # Не суперюзер - показ опубликованные
        if not (user.is_authenticated and user.is_staff):
            queryset = queryset.filter(status='published')
        
        return queryset

    def perform_create(self, serializer):
        # Автор - текущий пользователь
        serializer.save(author=self.request.user)

    # Тригер для email при публикации
    def perform_update(self, serializer):
        instance = serializer.save()
        if instance.status == 'published':
            from .tasks import send_event_email_task
            send_event_email_task.delay(
                event_id=instance.id,
                recipient_list=['admin@example.com'],
                subject=f"Опубликовано: {instance.title}",
                message=f"Мероприятие {instance.title} теперь доступно для всех."
            )

    # Экспорт
    @extend_schema(summary="Экспорт мероприятий в XLSX", tags=['Excel'])
    @action(detail=False, methods=['get'], url_path='export-xlsx')
    def export_xlsx(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Events"
        
        # Шапка
        columns = [
            'Название', 'Описание', 'Дата публикации', 
            'Дата начала', 'Дата завершения', 'Название места проведения', 
            'Гео-координаты', 'Рейтинг'
        ]
        sheet.append(columns)
        
        for event in queryset:
            coords = f"{event.location.lat}, {event.location.lon}"
            
            sheet.append([
                event.title,
                event.description,
                event.pub_date.replace(tzinfo=None) if event.pub_date else "",
                event.start_date.replace(tzinfo=None) if event.start_date else "",
                event.end_date.replace(tzinfo=None) if event.end_date else "",
                event.location.name,
                coords,
                event.rating
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=events_export.xlsx'
        workbook.save(response)
        return response
    
    # Импорт
    @extend_schema(
        summary="Импорт мероприятий из XLSX", 
        tags=['Excel'],
        request={'multipart/form-data': {'type': 'object', 'properties': {'file': {'type': 'string', 'format': 'binary'}}}}
    )
    @action(detail=False, methods=['post'], url_path='import-xlsx', permission_classes=[permissions.IsAdminUser])
    def import_xlsx(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "Файл не найден"}, status=400)
            
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue # Пропуск пустых строк
                
                title, desc, pub_dt, start_dt, end_dt, loc_name, loc_coords, rating = row
                
                # Парсим координаты "lat, lon"
                lat_raw, lon_raw = str(loc_coords).split(',')
                
                location, _ = Location.objects.get_or_create(
                    name=loc_name, 
                    defaults={'lat': float(lat_raw.strip()), 'lon': float(lon_raw.strip())}
                )
                
                # Создаем мероприятие
                Event.objects.create(
                    title=title,
                    description=desc,
                    start_date=start_dt,
                    end_date=end_dt,
                    location=location,
                    rating=rating if rating is not None else 0,
                    author=request.user,
                    status='published'
                )
                count += 1
                
            return Response({"status": f"Успешно импортировано {count} записей"})
        except Exception as e:
            return Response({"error": f"Ошибка при чтении файла: {str(e)}"}, status=400)